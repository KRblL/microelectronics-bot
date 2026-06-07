import io
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill


def split_grok_answer(text: str) -> tuple[str, str]:
    analogs_title = "Аналоги"
    analogs_index = text.find(analogs_title)
    if analogs_index == -1:
        return text.strip(), ""
    return text[:analogs_index].strip(), text[analogs_index:].strip()


def clean_value_with_url(value: str) -> str:
    return re.sub(r"\s*\(https?://[^)]+\)", "", value).strip()


def parse_component_block(block: str) -> dict[str, str]:
    lines = [line.strip() for line in block.splitlines() if line.strip()]
    data: dict[str, str] = {}
    in_characteristics = False
    for line in lines:
        if line in ("Основной компонент", "Аналоги"):
            continue
        if line.startswith("Производитель:"):
            data["Производитель"] = line.split(":", 1)[1].strip()
        elif line.startswith("Модель:"):
            data["Модель"] = line.split(":", 1)[1].strip()
        elif line.startswith("Год:"):
            data["Год"] = clean_value_with_url(line.split(":", 1)[1].strip())
        elif line.startswith("End of life:"):
            data["End of life"] = clean_value_with_url(line.split(":", 1)[1].strip())
        elif line.startswith("Тип компонента:"):
            data["Тип компонента"] = line.split(":", 1)[1].strip()
        elif line.startswith("Характеристики"):
            in_characteristics = True
        elif in_characteristics and line.startswith("•") and ":" in line:
            name, value = line.lstrip("•").strip().split(":", 1)
            data[name.strip()] = value.strip()
    return data


def parse_analogs_blocks(analogs_text: str) -> list[dict[str, str]]:
    text = analogs_text.strip()
    if text.startswith("Аналоги"):
        text = text[len("Аналоги") :].strip()
    blocks = []
    current = []
    for line in text.splitlines():
        if not line.strip():
            continue
        if line.startswith("Производитель:") and current:
            blocks.append("\n".join(current))
            current = [line]
        else:
            current.append(line)
    if current:
        blocks.append("\n".join(current))
    return [parse_component_block(block) for block in blocks]


def build_excel(base: dict[str, Any], analogs: list[dict[str, Any]]) -> tuple[bytes, str]:
    meta_keys = {"Производитель", "Модель", "Год", "Тип компонента", "End of life"}
    base_features = {key: value for key, value in base.items() if key not in meta_keys}
    analog_features = [{key: value for key, value in analog.items() if key not in meta_keys} for analog in analogs]
    common = set(base_features.keys())
    for item in analog_features:
        common = common.intersection(item.keys())
    features = sorted(common) if common else list(base_features.keys())

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Сравнение"
    header = ["Характеристика", "Основной компонент"] + [
        f"Аналог: {analog.get('Модель', f'#{index + 1}')}" for index, analog in enumerate(analogs)
    ]
    sheet.append(header)
    for column_index in range(1, len(header) + 1):
        sheet.column_dimensions[sheet.cell(row=1, column=column_index).column_letter].width = 32

    fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    for feature in features:
        row = [feature, base_features.get(feature, "")]
        for item in analog_features:
            row.append(item.get(feature, ""))
        sheet.append(row)
        row_index = sheet.max_row
        base_value = sheet.cell(row=row_index, column=2).value
        for column_index in range(3, len(header) + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            if cell.value and base_value and str(cell.value) != str(base_value):
                cell.fill = fill

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"{base.get('Модель', 'component')}_сравнение.xlsx"
    return buffer.read(), filename
